"""
Flower Shop Catalog - Backend API (ĐÃ FIX HOÀN TOÀN)
FastAPI + Supabase + Excel + PDF
"""

from fastapi import FastAPI, UploadFile, File, HTTPException, Form
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import JSONResponse, StreamingResponse
from supabase import create_client, Client
from pydantic import BaseModel
from typing import Optional, List
from PIL import Image
import os, io, uuid, traceback
from datetime import datetime
from dotenv import load_dotenv

# === THÊM IMPORT CHO EXCEL & PDF ===
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet

load_dotenv()

app = FastAPI(title="Flower Shop API")
app.add_middleware(CORSMiddleware, allow_origins=["*"], allow_credentials=True, allow_methods=["*"], allow_headers=["*"])

# Supabase
SUPABASE_URL = os.getenv("SUPABASE_URL")
SUPABASE_KEY = os.getenv("SUPABASE_KEY")
SUPABASE_BUCKET = os.getenv("SUPABASE_BUCKET", "flower-images")

try:
    supabase: Client = create_client(SUPABASE_URL, SUPABASE_KEY)
    print(f"Connected to Supabase: {SUPABASE_URL}")
except Exception as e:
    print(f"Supabase connection failed: {e}")
    supabase = None

# Models
class Flower(BaseModel):
    name: str; price: int; type: str; unit: str; image_url: Optional[str] = None; stock: Optional[int] = 0; tags: Optional[str] = ""

class FlowerUpdate(BaseModel):
    name: Optional[str] = None; price: Optional[int] = None; type: Optional[str] = None; unit: Optional[str] = None
    image_url: Optional[str] = None; stock: Optional[int] = None; tags: Optional[str] = None

class FlowerType(BaseModel): name: str; color: str = "#f5f5f5"
class UnitType(BaseModel): name: str

# Helper
def resize_image(image_data: bytes, max_size: tuple = (1200, 1200)) -> bytes:
    try:
        img = Image.open(io.BytesIO(image_data))
        if img.mode in ('RGBA', 'LA', 'P'):
            background = Image.new('RGB', img.size, (255, 255, 255))
            if img.mode == 'P': img = img.convert('RGBA')
            background.paste(img, mask=img.split()[-1] if img.mode == 'RGBA' else None)
            img = background
        img.thumbnail(max_size, Image.Resampling.LANCZOS)
        output = io.BytesIO()
        img.save(output, format='JPEG', quality=95)
        return output.getvalue()
    except Exception as e:
        print(f"Image resize error: {e}")
        raise

def upload_image_to_supabase(file: UploadFile, flower_name: str) -> str:
    try:
        if not supabase: raise Exception("Supabase not connected")
        image_data = file.file.read()
        resized_data = resize_image(image_data)
        filename = f"flower_{uuid.uuid4().hex}.jpg"
        supabase.storage.from_(SUPABASE_BUCKET).upload(filename, resized_data, file_options={"content-type": "image/jpeg", "upsert": "true"})
        url = supabase.storage.from_(SUPABASE_BUCKET).get_public_url(filename)
        url = str(url).strip()
        print(f"Image URL: {url}")
        return url
    except Exception as e:
        print(f"Image upload error: {e}")
        traceback.print_exc()
        return None

# === ENDPOINTS ===
@app.get("/") 
def root(): return {"message": "Flower Shop API", "version": "1.0", "supabase_connected": supabase is not None}

@app.get("/flowers")
def get_flowers(search: Optional[str] = None, type: Optional[str] = None, tags: Optional[str] = None, low_stock: bool = False, skip: int = 0, limit: int = 100):
    try:
        if not supabase: raise HTTPException(500, "DB not connected")
        query = supabase.table("flowers").select("*")
        if search: query = query.ilike("name", f"%{search}%")
        if type and type != "Tất cả": query = query.eq("type", type)
        if tags: query = query.ilike("tags", f"%{tags}%")
        if low_stock: query = query.lte("stock", 10)
        result = query.order("created_at", desc=True).range(skip, skip + limit - 1).execute()
        return {"flowers": result.data, "count": len(result.data)}
    except Exception as e: raise HTTPException(500, str(e))

@app.post("/flowers")
async def create_flower(name: str = Form(...), price: int = Form(...), type: str = Form(...), unit: str = Form(...), stock: int = Form(0), tags: str = Form(""), image: Optional[UploadFile] = File(None)):
    try:
        if not supabase: raise HTTPException(500, "DB not connected")
        image_url = upload_image_to_supabase(image, name) if image and image.filename else None
        data = {"name": name, "price": price, "type": type, "unit": unit, "stock": stock, "tags": tags, "image_url": image_url}
        result = supabase.table("flowers").insert(data).execute()
        return {"message": "Flower created", "flower": result.data[0]}
    except Exception as e: raise HTTPException(500, str(e))

@app.put("/flowers/{flower_id}")
async def update_flower(flower_id: str, name: Optional[str] = Form(None), price: Optional[int] = Form(None), type: Optional[str] = Form(None), unit: Optional[str] = Form(None), stock: Optional[int] = Form(None), tags: Optional[str] = Form(None), image: Optional[UploadFile] = File(None)):
    try:
        if not supabase: raise HTTPException(500, "DB not connected")
        data = {}
        if name: data["name"] = name
        if price: data["price"] = price
        if type: data["type"] = type
        if unit: data["unit"] = unit
        if stock is not None: data["stock"] = stock
        if tags is not None: data["tags"] = tags
        if image and image.filename:
            image_url = upload_image_to_supabase(image, name or "flower")
            if image_url: data["image_url"] = image_url
        if not data: raise HTTPException(400, "No data to update")
        result = supabase.table("flowers").update(data).eq("id", flower_id).execute()
        if not result.data: raise HTTPException(404, "Flower not found")
        return {"message": "Updated", "flower": result.data[0]}
    except HTTPException: raise
    except Exception as e: raise HTTPException(500, str(e))

@app.delete("/flowers/{flower_id}")
def delete_flower(flower_id: str):
    try:
        if not supabase: raise HTTPException(500, "DB not connected")
        flower = supabase.table("flowers").select("image_url").eq("id", flower_id).execute()
        if flower.data and flower.data[0].get("image_url"):
            filename = flower.data[0]["image_url"].split("/")[-1]
            try: supabase.storage.from_(SUPABASE_BUCKET).remove([filename])
            except: pass
        supabase.table("flowers").delete().eq("id", flower_id).execute()
        return {"message": "Deleted"}
    except Exception as e: raise HTTPException(500, str(e))

@app.get("/flower-types") 
def get_flower_types(): 
    result = supabase.table("flower_types").select("*").execute(); return {"types": result.data}

@app.get("/unit-types") 
def get_unit_types(): 
    result = supabase.table("unit_types").select("*").execute(); return {"units": result.data}

@app.get("/stats")
def get_stats():
    flowers = supabase.table("flowers").select("*").execute().data
    types = supabase.table("flower_types").select("*").execute().data
    total_value = sum(f["price"] for f in flowers)
    low_stock_count = sum(1 for f in flowers if f.get("stock", 0) <= 10)
    return {"total_flowers": len(flowers), "total_types": len(types), "total_value": total_value, "low_stock_count": low_stock_count}

# === EXCEL & PDF EXPORT ===
@app.get("/export/excel")
def export_excel(type: Optional[str] = None):
    try:
        query = supabase.table("flowers").select("*")
        if type and type != "Tất cả": query = query.eq("type", type)
        flowers = query.execute().data
        wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Catalog Hoa"
        headers = ["STT", "Tên hoa", "Giá (VNĐ)", "Loại", "Quy cách", "Tồn kho", "Tags"]
        header_fill = PatternFill(start_color="2e7d32", end_color="2e7d32", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        for col, h in enumerate(headers, 1):
            cell = ws.cell(1, col, h); cell.fill = header_fill; cell.font = header_font; cell.alignment = Alignment(horizontal="center")
        for idx, f in enumerate(flowers, 2):
            ws.cell(idx, 1, idx-1); ws.cell(idx, 2, f["name"]); ws.cell(idx, 3, f["price"])
            ws.cell(idx, 4, f["type"]); ws.cell(idx, 5, f["unit"]); ws.cell(idx, 6, f.get("stock", 0)); ws.cell(idx, 7, f.get("tags", ""))
        widths = [8, 35, 15, 15, 12, 10, 25]
        for i, w in enumerate(widths, 1): ws.column_dimensions[get_column_letter(i)].width = w
        output = io.BytesIO(); wb.save(output); output.seek(0)
        filename = f"Catalog_Hoa_{datetime.now():%Y%m%d_%H%M%S}.xlsx"
        return StreamingResponse(output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                 headers={"Content-Disposition": f"attachment; filename*=UTF-8''{filename}"})
    except Exception as e: raise HTTPException(500, str(e))

@app.get("/export/pdf")
def export_pdf(type: Optional[str] = None):
    try:
        query = supabase.table("flowers").select("*")
        if type and type != "Tất cả": query = query.eq("type", type)
        flowers = query.execute().data
        output = io.BytesIO(); doc = SimpleDocTemplate(output, pagesize=A4, topMargin=30)
        elements = []; styles = getSampleStyleSheet()
        elements.append(Paragraph("CATALOG HOA ĐẸP", styles['Heading1'])); elements.append(Spacer(1, 12))
        elements.append(Paragraph(f"Ngày xuất: {datetime.now():%d/%m/%Y %H:%M}", styles['Normal'])); elements.append(Spacer(1, 12))
        data = [["STT", "Tên hoa", "Giá", "Loại", "Quy cách", "Tồn", "Tags"]]
        for i, f in enumerate(flowers, 1):
            data.append([str(i), f["name"], f"{f['price']:,}".replace(",", "."), f["type"], f["unit"], str(f.get("stock", 0)), (f.get("tags", "") or "")[:30]])
        table = Table(data, colWidths=[30, 140, 70, 70, 60, 40, 80])
        table.setStyle(TableStyle([('BACKGROUND', (0,0), (-1,0), colors.HexColor("#2e7d32")), ('TEXTCOLOR', (0,0), (-1,0), colors.white),
            ('ALIGN', (0,0), (-1,-1), 'CENTER'), ('GRID', (0,0), (-1,-1), 0.5, colors.grey), ('FONTSIZE', (0,1), (-1,-1), 8)]))
        elements.append(table); elements.append(Spacer(1, 12)); elements.append(Paragraph(f"<b>Tổng: {len(flowers)} sản phẩm</b>", styles['Normal']))
        doc.build(elements); output.seek(0)
        filename = f"Catalog_Hoa_{datetime.now():%Y%m%d_%H%M%S}.pdf"
        return StreamingResponse(output, media_type="application/pdf", headers={"Content-Disposition": f"attachment; filename*=UTF-8''{filename}"})
    except Exception as e: raise HTTPException(500, str(e))

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)